import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]


def _workbook(path: Path) -> None:
    book = Workbook()
    sheets = [
        ("机构信息表", ["机构编号", "机构名称"], [["I1", "机构一"]]),
        ("指标清单表", ["指标编号", "指标名称", "指标含义", "指标单位"], [["M1", "指标一", "含义", "元"]]),
        ("衍生维度说明", ["衍生维度", "衍生口径说明"], [["同比", "当期与去年同期比较"]]),
        ("指标数据表", ["数据日期", "指标编号", "指标名称", "机构编号", "指标值"], [["2025-01-31", "M1", "指标一", "I1", "12.34"]]),
        ("问题答案清单", ["问题编号", "问题类型", "问题难度", "问题描述", "问题结果"], [["Q1", "单值", "简单", "虚构问题", "12.34"]]),
    ]
    first = book.active
    first.title = sheets[0][0]
    for index, (name, headers, rows) in enumerate(sheets):
        sheet = first if index == 0 else book.create_sheet(name)
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
    book.save(path)


class RealImportContractTest(unittest.TestCase):
    def test_import_rejects_formula_cells(self) -> None:
        from scripts.data.import_official_workbook import DataImportError, import_workbook

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory); source = root / "official.xlsx"; _workbook(source)
            book = Workbook()
            # Reuse a valid fixture, then mutate it through openpyxl.
            from openpyxl import load_workbook
            book = load_workbook(source); book["指标数据表"]["E2"] = "=1+1"; book.save(source)
            with self.assertRaises(DataImportError):
                import_workbook(source, root / "real", root / "private")

    def test_import_rejects_duplicate_fact_key(self) -> None:
        from scripts.data.import_official_workbook import DataImportError, import_workbook

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory); source = root / "official.xlsx"; _workbook(source)
            from openpyxl import load_workbook
            book = load_workbook(source); book["指标数据表"].append(["2025-01-31", "M1", "指标一", "I1", "12.34"]); book.save(source)
            with self.assertRaises(Exception):
                import_workbook(source, root / "real", root / "private")

    def test_import_publishes_matching_business_and_evaluation_manifests(self) -> None:
        from scripts.data.import_official_workbook import import_workbook

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "official.xlsx"
            _workbook(source)
            result = import_workbook(source, root / "real", root / "private")
            active = json.loads((root / "private" / "official" / "active_release.json").read_text(encoding="utf-8"))
            self.assertEqual(active["run_id"], result["run_id"])
            self.assertTrue(Path(active["business_database"]).is_file())
            self.assertTrue(Path(active["evaluation_database"]).is_file())


if __name__ == "__main__":
    unittest.main()
