from __future__ import annotations

import argparse
import sys
from collections import Counter
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from evaluation.models import EvalQuestion
from evaluation.question_bank import parse_question_id, write_questions

EXPECTED_HEADER = ["问题编号", "问题类型", "问题难度", "问题描述", "问题结果"]
EXPECTED_COUNTS = {"TRAIN": 120, "VAL": 40, "TST": 40}
SHEET_NAME = "问题答案清单"


def extract(
    source: Path,
    output: Path,
    sheet_name: str = SHEET_NAME,
    enforce_counts: bool = True,
) -> int:
    from openpyxl import load_workbook

    output = Path(output).resolve()
    if "private" not in output.parts:
        raise ValueError(
            f"输出路径必须位于受控 private 目录内，拒绝写入：{output}。"
        )
    workbook = load_workbook(Path(source), read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        header = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
        if header != EXPECTED_HEADER:
            raise ValueError(f"工作表表头不符合预期：{header}")
        questions: list[EvalQuestion] = []
        for raw in rows:
            if raw is None or raw[0] is None:
                continue
            question_id = str(raw[0]).strip()
            partition, difficulty = parse_question_id(question_id)
            questions.append(
                EvalQuestion(
                    question_id=question_id,
                    partition=partition,
                    difficulty=difficulty,
                    question_type=str(raw[1]).strip() if raw[1] is not None else "",
                    question_text=str(raw[3]).strip() if raw[3] is not None else "",
                    official_answer=str(raw[4]).strip() if raw[4] is not None else "",
                )
            )
    finally:
        workbook.close()

    counts = Counter(question.partition for question in questions)
    if enforce_counts and dict(counts) != EXPECTED_COUNTS:
        raise ValueError(f"分区题数与官方约定不符：{dict(counts)}，期望 {EXPECTED_COUNTS}。")

    write_questions(questions, output)
    return len(questions)


def main() -> int:
    parser = argparse.ArgumentParser(description="官方题库受控提取（输出进入 git 忽略目录）")
    parser.add_argument("--source", type=Path, required=True, help="官方 xlsx 路径（仓库外）")
    parser.add_argument(
        "--output",
        type=Path,
        default=ROOT / "data" / "private" / "evaluation" / "questions.jsonl",
    )
    parser.add_argument("--allow-any-counts", action="store_true")
    args = parser.parse_args()
    count = extract(args.source, args.output, enforce_counts=not args.allow_any_counts)
    print(f"EXTRACTED={count} -> {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
