from __future__ import annotations

import hashlib
import json
import os
import sqlite3
import uuid
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook


HEADERS = {
    "机构信息表": ("机构编号", "机构名称"),
    "指标清单表": ("指标编号", "指标名称", "指标含义", "指标单位"),
    "衍生维度说明": ("衍生维度", "衍生口径说明"),
    "指标数据表": ("数据日期", "指标编号", "指标名称", "机构编号", "指标值"),
    "问题答案清单": ("问题编号", "问题类型", "问题难度", "问题描述", "问题结果"),
}


class DataImportError(ValueError):
    pass


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _text(value: object, sheet: str, row: int) -> str:
    text = str(value).strip() if value is not None else ""
    if not text:
        raise DataImportError(f"{sheet} 第 {row} 行存在空字段")
    return text


def _date(value: object, row: int) -> str:
    if isinstance(value, datetime):
        if value.time().isoformat() != "00:00:00":
            raise DataImportError(f"指标数据表 第 {row} 行日期包含时间")
        value = value.date()
    if isinstance(value, date):
        return value.isoformat()
    try:
        return datetime.strptime(str(value).strip(), "%Y-%m-%d").date().isoformat()
    except ValueError as exc:
        raise DataImportError(f"指标数据表 第 {row} 行日期无效") from exc


def _decimal(value: object, row: int) -> Decimal:
    try:
        number = Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise DataImportError(f"指标数据表 第 {row} 行指标值无效") from exc
    if not number.is_finite():
        raise DataImportError(f"指标数据表 第 {row} 行指标值无效")
    return number


def _rows(book, name: str):
    sheet = book[name]
    headers = tuple(cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1)))
    if len(headers) != len(set(headers)) or set(headers) != set(HEADERS[name]):
        raise DataImportError(f"{name} 表头不符合输入契约")
    positions = {header: index for index, header in enumerate(headers)}
    for row_number, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        if any(cell.data_type == "f" for cell in row):
            raise DataImportError(f"{name} 第 {row_number} 行包含公式")
        values = {header: row[positions[header]].value for header in HEADERS[name]}
        yield row_number, values


def import_workbook(source: Path, real_root: Path, private_root: Path) -> dict[str, str]:
    source = Path(source).resolve()
    if not source.is_file():
        raise FileNotFoundError(source)
    private_root = Path(private_root)
    private_root.mkdir(parents=True, exist_ok=True)
    staging_path = private_root / f".staging-{uuid.uuid4().hex}.db"
    staging = sqlite3.connect(staging_path)
    staging.execute("CREATE TABLE facts (data_date TEXT NOT NULL, metric_id TEXT NOT NULL, institution_id TEXT NOT NULL, raw_value TEXT NOT NULL, row_number INTEGER NOT NULL, PRIMARY KEY(institution_id, metric_id, data_date))")
    book = load_workbook(source, read_only=True, data_only=False)
    try:
        missing = set(HEADERS).difference(book.sheetnames)
        if missing:
            raise DataImportError(f"缺少工作表: {sorted(missing)}")
        institutions = [( _text(v["机构编号"], "机构信息表", row), _text(v["机构名称"], "机构信息表", row)) for row, v in _rows(book, "机构信息表")]
        metrics = [( _text(v["指标编号"], "指标清单表", row), _text(v["指标名称"], "指标清单表", row), _text(v["指标含义"], "指标清单表", row), _text(v["指标单位"], "指标清单表", row)) for row, v in _rows(book, "指标清单表")]
        metric_names = {item[0]: item[1] for item in metrics}
        institution_ids = {item[0] for item in institutions}
        scales: dict[str, int] = {}
        batch = []
        for row, values in _rows(book, "指标数据表"):
            metric_id = _text(values["指标编号"], "指标数据表", row)
            institution_id = _text(values["机构编号"], "指标数据表", row)
            if metric_id not in metric_names or institution_id not in institution_ids or metric_names[metric_id] != _text(values["指标名称"], "指标数据表", row):
                raise DataImportError(f"指标数据表 第 {row} 行关联或名称不一致")
            number = _decimal(values["指标值"], row)
            scales[metric_id] = max(scales.get(metric_id, 0), max(0, -number.as_tuple().exponent))
            batch.append((_date(values["数据日期"], row), metric_id, institution_id, str(number), row))
            if len(batch) == 1000:
                staging.executemany("INSERT INTO facts VALUES (?, ?, ?, ?, ?)", batch)
                batch.clear()
        if batch:
            staging.executemany("INSERT INTO facts VALUES (?, ?, ?, ?, ?)", batch)
        staging.commit()
        fact_count = staging.execute("SELECT COUNT(*) FROM facts").fetchone()[0]
        derived = [(_text(v["衍生维度"], "衍生维度说明", row), _text(v["衍生口径说明"], "衍生维度说明", row)) for row, v in _rows(book, "衍生维度说明")]
        questions = [tuple(_text(v[key], "问题答案清单", row) for key in HEADERS["问题答案清单"]) for row, v in _rows(book, "问题答案清单")]
    except Exception:
        staging.close()
        staging_path.unlink(missing_ok=True)
        raise
    finally:
        book.close()
    run_id, source_sha256 = uuid.uuid4().hex, _sha256(source)
    real_dir, evaluation_dir = Path(real_root) / "releases" / run_id, Path(private_root) / "evaluation" / "releases" / run_id
    real_dir.mkdir(parents=True, exist_ok=False); evaluation_dir.mkdir(parents=True, exist_ok=False)
    business, evaluation = real_dir / "bankinsight_real.db", evaluation_dir / "questions.db"
    schema = Path(__file__).resolve().parents[2] / "sql" / "real_schema.sql"
    now = datetime.now(timezone.utc).isoformat()
    def manifest(conn, table_counts):
        conn.execute("INSERT INTO import_manifest VALUES (?, ?, '1', ?, ?, ?, ?, ?)", (run_id, source_sha256, now, *table_counts))
    conn = sqlite3.connect(business)
    try:
        conn.execute("PRAGMA foreign_keys=ON"); conn.executescript(schema.read_text(encoding="utf-8"))
        conn.executemany("INSERT INTO institutions VALUES (?, ?)", institutions)
        conn.executemany("INSERT INTO metrics VALUES (?, ?, ?, ?, ?)", [(a,b,c,d,scales[a]) for a,b,c,d in metrics])
        cursor = staging.execute("SELECT data_date, metric_id, institution_id, raw_value FROM facts ORDER BY row_number")
        while rows := cursor.fetchmany(1000):
            conn.executemany("INSERT INTO metric_facts VALUES (?, ?, ?, ?)", [(d,m,i,int(Decimal(value) * (10 ** scales[m]))) for d,m,i,value in rows])
        conn.executemany("INSERT INTO derived_dimensions VALUES (?, ?)", derived)
        manifest(conn, (len(institutions), len(metrics), fact_count, len(derived)))
        if conn.execute("PRAGMA foreign_key_check").fetchall(): raise DataImportError("外键检查失败")
        conn.commit()
    finally:
        conn.close()
    conn = sqlite3.connect(evaluation)
    try:
        conn.execute("CREATE TABLE import_manifest (run_id TEXT PRIMARY KEY, source_sha256 TEXT NOT NULL, schema_version TEXT NOT NULL, created_at_utc TEXT NOT NULL, institution_count INTEGER NOT NULL, metric_count INTEGER NOT NULL, fact_count INTEGER NOT NULL, derived_dimension_count INTEGER NOT NULL)")
        conn.execute("CREATE TABLE evaluation_questions (question_id TEXT PRIMARY KEY, question_type TEXT NOT NULL, question_difficulty TEXT NOT NULL, question_text TEXT NOT NULL, expected_result TEXT NOT NULL)")
        conn.executemany("INSERT INTO evaluation_questions VALUES (?, ?, ?, ?, ?)", questions); manifest(conn, (len(institutions), len(metrics), fact_count, len(derived)))
        conn.commit()
    finally:
        conn.close()
    staging.close(); staging_path.unlink(missing_ok=True)
    active_dir = private_root / "official"; active_dir.mkdir(parents=True, exist_ok=True)
    active = {"run_id": run_id, "source_sha256": source_sha256, "business_database": str(business.resolve()), "evaluation_database": str(evaluation.resolve())}
    temporary = active_dir / "active_release.json.tmp"; temporary.write_text(json.dumps(active, ensure_ascii=False, indent=2), encoding="utf-8"); os.replace(temporary, active_dir / "active_release.json")
    return active
