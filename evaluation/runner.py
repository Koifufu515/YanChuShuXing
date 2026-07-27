from __future__ import annotations

import argparse
import hashlib
import http.client
import json
import os
import sqlite3
import statistics
import subprocess
import sys
import time
import uuid
from dataclasses import asdict, is_dataclass
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any, Callable, Iterable, Mapping, Protocol, Sequence

from openpyxl import load_workbook

from evaluation.models import (
    EvaluationPaths,
    EvaluationQuestion,
    VALID_SPLITS,
)


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = (
    PROJECT_ROOT
    / "data"
    / "private"
    / "evaluation"
    / "source"
    / "问题答案清单.xlsx"
)
DEFAULT_RUNS_ROOT = (
    PROJECT_ROOT
    / "data"
    / "private"
    / "evaluation"
    / "runs"
)

REQUIRED_HEADERS = (
    "问题编号",
    "问题类型",
    "问题难度",
    "问题描述",
    "问题结果",
)
CORRECTED_HEADER = "更正后答案"

STRUCTURED_NON_EXECUTABLE_CODES = frozenset(
    {
        "CLARIFICATION_REQUIRED",
        "PENDING_PROJECT_DEFINITION",
        "DATA_UNAVAILABLE",
    }
)
TRANSIENT_ERROR_CODES = frozenset(
    {
        "LLM_UNAVAILABLE",
        "LLM_TIMEOUT",
        "DATABASE_UNAVAILABLE",
        "QUERY_TIMEOUT",
    }
)

TRANSIENT_EXCEPTION_NAMES = frozenset(
    {
        "IncompleteRead",
        "RemoteDisconnected",
        "ProtocolError",
        "ReadTimeoutError",
        "ConnectTimeoutError",
        "MaxRetryError",
    }
)


def is_retryable_exception(exception: BaseException) -> bool:
    """识别HTTP连接中断、超时和连接重置等瞬时异常。"""
    visited: set[int] = set()
    current: BaseException | None = exception
    while current is not None and id(current) not in visited:
        visited.add(id(current))
        if isinstance(
            current,
            (
                http.client.IncompleteRead,
                http.client.RemoteDisconnected,
                TimeoutError,
                ConnectionResetError,
                ConnectionAbortedError,
                BrokenPipeError,
            ),
        ):
            return True
        if type(current).__name__ in TRANSIENT_EXCEPTION_NAMES:
            return True
        next_exception = current.__cause__ or current.__context__
        current = (
            next_exception
            if isinstance(next_exception, BaseException)
            else None
        )
    return False


class QueryServiceLike(Protocol):
    def run(self, command: Any) -> Any:
        ...


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def normalized_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def write_json_atomic(path: Path, payload: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(
        json.dumps(
            to_jsonable(payload),
            ensure_ascii=False,
            indent=2,
            sort_keys=True,
        ),
        encoding="utf-8",
    )
    os.replace(temporary, path)


def append_jsonl(path: Path, payload: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    line = json.dumps(
        to_jsonable(payload),
        ensure_ascii=False,
        separators=(",", ":"),
        sort_keys=True,
    )
    with path.open("a", encoding="utf-8") as handle:
        handle.write(line + "\n")
        handle.flush()
        os.fsync(handle.fileno())


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    if not path.is_file():
        return []

    records: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as handle:
        for line_number, raw_line in enumerate(handle, start=1):
            line = raw_line.strip()
            if not line:
                continue
            try:
                payload = json.loads(line)
            except json.JSONDecodeError as exc:
                raise RuntimeError(
                    f"{path} 第{line_number}行不是合法JSON。"
                ) from exc
            if not isinstance(payload, dict):
                raise RuntimeError(
                    f"{path} 第{line_number}行顶层必须是JSON对象。"
                )
            records.append(payload)
    return records


def latest_records_by_question(
    records: Iterable[Mapping[str, Any]],
) -> dict[str, dict[str, Any]]:
    latest: dict[str, dict[str, Any]] = {}
    for record in records:
        question_id = record.get("question_id")
        if isinstance(question_id, str) and question_id:
            latest[question_id] = dict(record)
    return latest


def to_jsonable(value: object) -> object:
    if is_dataclass(value):
        return to_jsonable(asdict(value))
    if isinstance(value, Mapping):
        return {
            str(key): to_jsonable(item)
            for key, item in value.items()
        }
    if isinstance(value, (list, tuple, set, frozenset)):
        return [to_jsonable(item) for item in value]
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Path):
        return str(value)
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if hasattr(value, "__dict__"):
        return to_jsonable(vars(value))
    return str(value)


def load_questions(source: Path) -> list[EvaluationQuestion]:
    source = Path(source)
    if not source.is_file():
        raise FileNotFoundError(f"题库文件不存在：{source}")

    workbook = load_workbook(
        source,
        read_only=True,
        data_only=True,
    )
    try:
        if "问题答案清单" not in workbook.sheetnames:
            raise ValueError("题库缺少“问题答案清单”工作表。")

        sheet = workbook["问题答案清单"]
        raw_headers = [
            normalized_text(cell.value)
            for cell in next(sheet.iter_rows(min_row=1, max_row=1))
        ]
        if len(raw_headers) != len(set(raw_headers)):
            raise ValueError("问题答案清单存在重复表头。")

        positions = {
            header: index
            for index, header in enumerate(raw_headers)
            if header
        }
        missing = [
            header
            for header in REQUIRED_HEADERS
            if header not in positions
        ]
        if missing:
            raise ValueError(
                f"问题答案清单缺少必要列：{missing}"
            )

        questions: list[EvaluationQuestion] = []
        seen_ids: set[str] = set()

        for row_number, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            question_id = normalized_text(
                _row_value(row, positions["问题编号"])
            )
            if not question_id:
                if all(normalized_text(item) == "" for item in row):
                    continue
                raise ValueError(
                    f"问题答案清单第{row_number}行缺少问题编号。"
                )
            if question_id in seen_ids:
                raise ValueError(f"问题编号重复：{question_id}")
            seen_ids.add(question_id)

            split = question_id.split("-", 1)[0].upper()
            if split == "TEST":
                split = "TST"
            if split not in VALID_SPLITS:
                raise ValueError(
                    f"问题编号无法识别数据分区：{question_id}"
                )

            official_answer = normalized_text(
                _row_value(row, positions["问题结果"])
            )
            corrected_answer = None
            corrected_position = positions.get(CORRECTED_HEADER)
            if corrected_position is not None:
                corrected_text = normalized_text(
                    _row_value(row, corrected_position)
                )
                corrected_answer = corrected_text or None

            expected_answer = corrected_answer or official_answer
            if not expected_answer:
                raise ValueError(
                    f"问题{question_id}缺少可用预期答案。"
                )

            question_text = normalized_text(
                _row_value(row, positions["问题描述"])
            )
            if not question_text:
                raise ValueError(
                    f"问题{question_id}缺少问题描述。"
                )

            questions.append(
                EvaluationQuestion(
                    question_id=question_id,
                    split=split,
                    question_type=normalized_text(
                        _row_value(row, positions["问题类型"])
                    ),
                    difficulty=normalized_text(
                        _row_value(row, positions["问题难度"])
                    ),
                    question=question_text,
                    official_answer=official_answer,
                    corrected_answer=corrected_answer,
                    expected_answer=expected_answer,
                    answer_source=(
                        "corrected"
                        if corrected_answer is not None
                        else "official"
                    ),
                )
            )
    finally:
        workbook.close()

    return questions


def _row_value(row: Sequence[object], position: int) -> object:
    return row[position] if position < len(row) else None


def select_questions(
    questions: Sequence[EvaluationQuestion],
    split: str,
    limit: int | None = None,
    question_ids: Sequence[str] | None = None,
) -> list[EvaluationQuestion]:
    normalized_split = split.upper()
    if normalized_split not in VALID_SPLITS:
        raise ValueError(f"不支持的数据分区：{split}")

    selected = [
        question
        for question in questions
        if question.split == normalized_split
    ]

    if question_ids:
        requested = set(question_ids)
        selected = [
            question
            for question in selected
            if question.question_id in requested
        ]
        found = {question.question_id for question in selected}
        missing = sorted(requested - found)
        if missing:
            raise ValueError(
                f"指定问题不属于{normalized_split}或不存在：{missing}"
            )

    if limit is not None:
        if limit < 1:
            raise ValueError("--limit必须是正整数。")
        selected = selected[:limit]

    return selected


def classify_outcome(
    error: Mapping[str, Any] | None,
    metadata: Mapping[str, Any] | None,
) -> tuple[str, str | None]:
    if error is None:
        return "success", None

    code = str(error.get("code") or "UNKNOWN_ERROR")
    if code in STRUCTURED_NON_EXECUTABLE_CODES:
        return "non_executable", "status_resolution"

    failure_reason = (
        metadata.get("failure_reason")
        if isinstance(metadata, Mapping)
        else None
    )
    if code == "LLM_PROVIDER_ERROR" and failure_reason == "invalid_query_plan":
        return "failed", "query_plan_validation"
    if code.startswith("LLM_") or code in {
        "INVALID_SEMANTIC_OUTPUT",
        "INVALID_SQL_OUTPUT",
    }:
        return "failed", "query_planner"
    if code in {
        "QUERY_EXECUTION_ERROR",
        "QUERY_TIMEOUT",
        "DATABASE_UNAVAILABLE",
    }:
        return "failed", "query_execution"
    if code == "CONFIGURATION_ERROR":
        return "failed", "configuration"
    return "failed", "pipeline"


def outcome_record(
    question: EvaluationQuestion,
    outcome: object,
    run_id: str,
    sequence: int,
    revision: int,
    started_at: str,
    finished_at: str,
    duration_ms: float,
    attempt_count: int,
) -> dict[str, Any]:
    error_value = getattr(outcome, "error", None)
    metadata_value = getattr(outcome, "metadata", None)
    error = to_jsonable(error_value)
    metadata = to_jsonable(metadata_value)

    error_mapping = error if isinstance(error, dict) else None
    metadata_mapping = (
        metadata if isinstance(metadata, dict) else None
    )
    run_status, failure_stage = classify_outcome(
        error_mapping,
        metadata_mapping,
    )

    query_plan = (
        metadata_mapping.get("query_plan")
        if metadata_mapping
        else None
    )
    plan_status = None
    if isinstance(query_plan, dict):
        status = query_plan.get("status")
        if isinstance(status, dict):
            plan_status = status.get("code")

    return {
        "run_id": run_id,
        "sequence": sequence,
        "revision": revision,
        "question_id": question.question_id,
        "split": question.split,
        "question_type": question.question_type,
        "difficulty": question.difficulty,
        "question": question.question,
        "official_answer": question.official_answer,
        "corrected_answer": question.corrected_answer,
        "expected_answer": question.expected_answer,
        "answer_source": question.answer_source,
        "comparison_status": "not_scored",
        "run_status": run_status,
        "failure_stage": failure_stage,
        "started_at_utc": started_at,
        "finished_at_utc": finished_at,
        "duration_ms": round(duration_ms, 3),
        "attempt_count": attempt_count,
        "request_id": getattr(outcome, "request_id", None),
        "columns": to_jsonable(
            getattr(outcome, "columns", [])
        ),
        "rows": to_jsonable(getattr(outcome, "rows", [])),
        "summary": getattr(outcome, "summary", None),
        "warnings": to_jsonable(
            getattr(outcome, "warnings", [])
        ),
        "error": error_mapping,
        "metadata": metadata_mapping,
        "plan_status": plan_status,
        "plan_repair_attempted": (
            metadata_mapping.get("plan_repair_attempted")
            if metadata_mapping
            else None
        ),
    }


def exception_record(
    question: EvaluationQuestion,
    run_id: str,
    sequence: int,
    revision: int,
    started_at: str,
    finished_at: str,
    duration_ms: float,
    attempt_count: int,
    exception: BaseException,
) -> dict[str, Any]:
    return {
        "run_id": run_id,
        "sequence": sequence,
        "revision": revision,
        "question_id": question.question_id,
        "split": question.split,
        "question_type": question.question_type,
        "difficulty": question.difficulty,
        "question": question.question,
        "official_answer": question.official_answer,
        "corrected_answer": question.corrected_answer,
        "expected_answer": question.expected_answer,
        "answer_source": question.answer_source,
        "comparison_status": "not_scored",
        "run_status": "exception",
        "failure_stage": "runner_exception",
        "started_at_utc": started_at,
        "finished_at_utc": finished_at,
        "duration_ms": round(duration_ms, 3),
        "attempt_count": attempt_count,
        "request_id": None,
        "columns": [],
        "rows": [],
        "summary": None,
        "warnings": [],
        "error": {
            "code": "RUNNER_EXCEPTION",
            "message": str(exception)[:500],
            "retryable": is_retryable_exception(exception),
            "exception_type": type(exception).__name__,
        },
        "metadata": None,
        "plan_status": None,
        "plan_repair_attempted": None,
    }


def execute_batch(
    pipeline: QueryServiceLike,
    questions: Sequence[EvaluationQuestion],
    paths: EvaluationPaths,
    manifest: Mapping[str, Any],
    *,
    resume: bool = True,
    rerun_failures: bool = False,
    retries: int = 0,
    delay_seconds: float = 0.0,
    command_factory: Callable[[str, str, str, int], object] | None = None,
) -> dict[str, Any]:
    if retries < 0:
        raise ValueError("--retries不能为负数。")
    if delay_seconds < 0:
        raise ValueError("--delay-seconds不能为负数。")

    paths.run_root.mkdir(parents=True, exist_ok=True)
    _prepare_manifest(paths.manifest, manifest, resume)
    resolved_command_factory = command_factory or _query_command

    all_records = read_jsonl(paths.results)
    latest = latest_records_by_question(all_records)
    revisions = _revision_counts(all_records)

    skipped = 0
    executed = 0
    interrupted = False

    try:
        for sequence, question in enumerate(questions, start=1):
            previous = latest.get(question.question_id)
            if previous is not None and _should_skip(
                previous,
                rerun_failures=rerun_failures,
            ):
                skipped += 1
                print(
                    f"[{sequence}/{len(questions)}] "
                    f"{question.question_id} skipped",
                    flush=True,
                )
                continue

            revision = revisions.get(question.question_id, 0) + 1
            started_at = utc_now()
            started = time.perf_counter()
            attempt_count = 0
            final_outcome: object | None = None
            final_exception: BaseException | None = None

            for attempt in range(retries + 1):
                attempt_count = attempt + 1
                final_exception = None
                try:
                    command = resolved_command_factory(
                        question.question,
                        str(manifest["run_id"]),
                        question.question_id,
                        attempt_count,
                    )
                    final_outcome = pipeline.run(command)
                except KeyboardInterrupt:
                    raise
                except Exception as exc:
                    final_outcome = None
                    final_exception = exc

                if final_exception is not None:
                    if (
                        is_retryable_exception(final_exception)
                        and attempt < retries
                    ):
                        if delay_seconds:
                            time.sleep(delay_seconds)
                        continue
                    break

                error = to_jsonable(
                    getattr(final_outcome, "error", None)
                )
                error_code = (
                    error.get("code")
                    if isinstance(error, dict)
                    else None
                )
                if (
                    error_code in TRANSIENT_ERROR_CODES
                    and attempt < retries
                ):
                    continue
                break

            finished_at = utc_now()
            duration_ms = (
                time.perf_counter() - started
            ) * 1000

            if final_exception is not None:
                record = exception_record(
                    question=question,
                    run_id=str(manifest["run_id"]),
                    sequence=sequence,
                    revision=revision,
                    started_at=started_at,
                    finished_at=finished_at,
                    duration_ms=duration_ms,
                    attempt_count=attempt_count,
                    exception=final_exception,
                )
            else:
                record = outcome_record(
                    question=question,
                    outcome=final_outcome,
                    run_id=str(manifest["run_id"]),
                    sequence=sequence,
                    revision=revision,
                    started_at=started_at,
                    finished_at=finished_at,
                    duration_ms=duration_ms,
                    attempt_count=attempt_count,
                )

            append_jsonl(paths.results, record)
            all_records.append(record)
            latest[question.question_id] = record
            revisions[question.question_id] = revision
            executed += 1

            error_code = (
                record.get("error", {}).get("code")
                if isinstance(record.get("error"), dict)
                else None
            )
            detail = (
                f" error={error_code}"
                if error_code
                else ""
            )
            print(
                f"[{sequence}/{len(questions)}] "
                f"{question.question_id} "
                f"{record['run_status']} "
                f"{record['duration_ms']}ms{detail}",
                flush=True,
            )

            _write_current_reports(
                paths=paths,
                selected_questions=questions,
                all_records=all_records,
                skipped=skipped,
                executed=executed,
                interrupted=False,
            )

            if delay_seconds:
                time.sleep(delay_seconds)
    except KeyboardInterrupt:
        interrupted = True
    finally:
        summary = _write_current_reports(
            paths=paths,
            selected_questions=questions,
            all_records=all_records,
            skipped=skipped,
            executed=executed,
            interrupted=interrupted,
        )

    if interrupted:
        raise KeyboardInterrupt

    return summary


def _query_command(
    question: str,
    run_id: str,
    question_id: str,
    attempt: int,
) -> object:
    backend_root = PROJECT_ROOT / "backend"
    if str(backend_root) not in sys.path:
        sys.path.insert(0, str(backend_root))

    from app.application.models import QueryCommand

    return QueryCommand(
        question=question,
        user_id="batch_evaluation",
        conversation_id=run_id,
        request_id=(
            f"eval-{question_id.lower()}-"
            f"{attempt}-{uuid.uuid4().hex[:12]}"
        ),
    )


def _prepare_manifest(
    manifest_path: Path,
    manifest: Mapping[str, Any],
    resume: bool,
) -> None:
    if manifest_path.is_file():
        if not resume:
            raise RuntimeError(
                f"运行目录已存在：{manifest_path.parent}。"
            )
        existing = json.loads(
            manifest_path.read_text(encoding="utf-8")
        )
        _validate_resume_manifest(existing, manifest)
        return

    if manifest_path.parent.exists() and not resume:
        existing_items = list(manifest_path.parent.iterdir())
        if existing_items:
            raise RuntimeError(
                f"运行目录非空：{manifest_path.parent}。"
            )
    write_json_atomic(manifest_path, manifest)


def _validate_resume_manifest(
    existing: object,
    current: Mapping[str, Any],
) -> None:
    if not isinstance(existing, dict):
        raise RuntimeError("已有manifest.json格式不正确。")

    comparable_paths = (
        ("run_id",),
        ("split",),
        ("source", "sha256"),
        ("selected_question_ids",),
        ("system", "code_commit"),
        ("system", "runner_sha256"),
        ("system", "prompt_sha256"),
        ("system", "schema_sha256"),
        ("system", "context_sha256"),
        ("database", "run_id"),
        ("database", "source_sha256"),
    )
    mismatches: list[str] = []
    for path in comparable_paths:
        old_value = _nested(existing, path)
        new_value = _nested(current, path)
        if old_value != new_value:
            mismatches.append(".".join(path))

    if mismatches:
        raise RuntimeError(
            "当前环境与已有运行记录不一致，"
            f"请使用新的--run-id。差异字段：{mismatches}"
        )


def _nested(payload: Mapping[str, Any], path: Sequence[str]) -> object:
    value: object = payload
    for part in path:
        if not isinstance(value, Mapping):
            return None
        value = value.get(part)
    return value


def _revision_counts(
    records: Iterable[Mapping[str, Any]],
) -> dict[str, int]:
    counts: dict[str, int] = {}
    for record in records:
        question_id = record.get("question_id")
        revision = record.get("revision")
        if isinstance(question_id, str):
            value = revision if isinstance(revision, int) else 1
            counts[question_id] = max(
                counts.get(question_id, 0),
                value,
            )
    return counts


def _should_skip(
    previous: Mapping[str, Any],
    rerun_failures: bool,
) -> bool:
    if not rerun_failures:
        return True
    return previous.get("run_status") in {
        "success",
        "non_executable",
    }


def _write_current_reports(
    paths: EvaluationPaths,
    selected_questions: Sequence[EvaluationQuestion],
    all_records: Sequence[Mapping[str, Any]],
    skipped: int,
    executed: int,
    interrupted: bool,
) -> dict[str, Any]:
    selected_ids = {
        question.question_id for question in selected_questions
    }
    latest = latest_records_by_question(
        record
        for record in all_records
        if record.get("question_id") in selected_ids
    )
    ordered = [
        latest[question.question_id]
        for question in selected_questions
        if question.question_id in latest
    ]

    summary = build_summary(
        selected_questions=selected_questions,
        records=ordered,
        skipped=skipped,
        executed=executed,
        interrupted=interrupted,
    )
    write_json_atomic(paths.summary, summary)

    failures = [
        record
        for record in ordered
        if record.get("run_status") != "success"
    ]
    _write_jsonl_snapshot(paths.failures, failures)
    return summary


def _write_jsonl_snapshot(
    path: Path,
    records: Sequence[Mapping[str, Any]],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    with temporary.open("w", encoding="utf-8") as handle:
        for record in records:
            handle.write(
                json.dumps(
                    to_jsonable(record),
                    ensure_ascii=False,
                    separators=(",", ":"),
                    sort_keys=True,
                )
                + "\n"
            )
    os.replace(temporary, path)


def build_summary(
    selected_questions: Sequence[EvaluationQuestion],
    records: Sequence[Mapping[str, Any]],
    skipped: int,
    executed: int,
    interrupted: bool,
) -> dict[str, Any]:
    status_counts: dict[str, int] = {}
    error_counts: dict[str, int] = {}
    plan_status_counts: dict[str, int] = {}
    durations: list[float] = []
    repaired_total = 0

    for record in records:
        run_status = str(record.get("run_status") or "unknown")
        status_counts[run_status] = (
            status_counts.get(run_status, 0) + 1
        )

        duration = record.get("duration_ms")
        if isinstance(duration, (int, float)):
            durations.append(float(duration))

        error = record.get("error")
        if isinstance(error, Mapping):
            code = str(error.get("code") or "UNKNOWN_ERROR")
            error_counts[code] = error_counts.get(code, 0) + 1

        plan_status = record.get("plan_status")
        if isinstance(plan_status, str) and plan_status:
            plan_status_counts[plan_status] = (
                plan_status_counts.get(plan_status, 0) + 1
            )

        if record.get("plan_repair_attempted") is True:
            repaired_total += 1

    completed_total = len(records)
    return {
        "updated_at_utc": utc_now(),
        "interrupted": interrupted,
        "selected_total": len(selected_questions),
        "completed_total": completed_total,
        "remaining_total": max(
            0,
            len(selected_questions) - completed_total,
        ),
        "executed_this_invocation": executed,
        "skipped_this_invocation": skipped,
        "run_status_counts": dict(sorted(status_counts.items())),
        "error_code_counts": dict(sorted(error_counts.items())),
        "plan_status_counts": dict(
            sorted(plan_status_counts.items())
        ),
        "plan_repair_attempted_total": repaired_total,
        "duration_ms": {
            "mean": (
                round(statistics.fmean(durations), 3)
                if durations
                else None
            ),
            "p50": _percentile(durations, 0.50),
            "p95": _percentile(durations, 0.95),
            "max": round(max(durations), 3) if durations else None,
        },
        "comparison_status": "not_scored",
    }


def _percentile(values: Sequence[float], fraction: float) -> float | None:
    if not values:
        return None
    ordered = sorted(values)
    if len(ordered) == 1:
        return round(ordered[0], 3)
    position = (len(ordered) - 1) * fraction
    lower = int(position)
    upper = min(lower + 1, len(ordered) - 1)
    weight = position - lower
    result = (
        ordered[lower] * (1 - weight)
        + ordered[upper] * weight
    )
    return round(result, 3)


def build_runtime() -> tuple[QueryServiceLike, dict[str, Any]]:
    backend_root = PROJECT_ROOT / "backend"
    if str(backend_root) not in sys.path:
        sys.path.insert(0, str(backend_root))

    from app.bootstrap.container import build_pipeline
    from app.core.data_source import resolve_database_path
    from app.core.settings import Settings

    settings = Settings.from_env(PROJECT_ROOT / ".env")
    database_path = resolve_database_path(
        PROJECT_ROOT,
        settings.data_environment,
        settings.database_path_override,
    )
    pipeline = build_pipeline(
        database_path=database_path,
        settings=settings,
    )
    runtime = {
        "settings": settings,
        "database_path": database_path,
    }
    return pipeline, runtime


def build_manifest(
    *,
    run_id: str,
    split: str,
    source: Path,
    selected_questions: Sequence[EvaluationQuestion],
    runtime: Mapping[str, Any],
) -> dict[str, Any]:
    settings = runtime["settings"]
    database_path = Path(runtime["database_path"])

    config_dir = PROJECT_ROOT / "config" / "query_planner"
    prompt_path = config_dir / "query_planner_prompt.md"
    schema_path = config_dir / "query_plan.schema.json"
    context_path = config_dir / "query_planner_context.json"

    return {
        "manifest_version": 1,
        "run_id": run_id,
        "created_at_utc": utc_now(),
        "split": split,
        "selected_total": len(selected_questions),
        "selected_question_ids": [
            question.question_id for question in selected_questions
        ],
        "source": {
            "path": str(source.resolve()),
            "sha256": sha256_file(source),
        },
        "system": {
            "code_commit": _git_output(
                ["rev-parse", "HEAD"],
                default="unknown",
            ),
            "git_dirty": bool(
                _git_output(
                    ["status", "--porcelain"],
                    default="",
                )
            ),
            "runner_sha256": sha256_file(Path(__file__)),
            "prompt_sha256": sha256_file(prompt_path),
            "schema_sha256": sha256_file(schema_path),
            "context_sha256": sha256_file(context_path),
            "data_environment": settings.data_environment,
            "provider": settings.llm_provider,
            "model": settings.llm_model,
            "timeout_seconds": settings.llm_timeout_seconds,
            "temperature": settings.llm_temperature,
        },
        "database": read_database_manifest(database_path),
    }


def read_database_manifest(database_path: Path) -> dict[str, Any]:
    connection = sqlite3.connect(
        f"file:{database_path.resolve()}?mode=ro",
        uri=True,
    )
    try:
        row = connection.execute(
            "SELECT run_id, source_sha256, schema_version, "
            "created_at_utc, institution_count, metric_count, "
            "fact_count, derived_dimension_count "
            "FROM import_manifest"
        ).fetchone()
    finally:
        connection.close()

    if row is None:
        raise RuntimeError("正式数据库缺少import_manifest。")

    return {
        "path": str(database_path.resolve()),
        "run_id": row[0],
        "source_sha256": row[1],
        "schema_version": row[2],
        "created_at_utc": row[3],
        "institution_count": row[4],
        "metric_count": row[5],
        "fact_count": row[6],
        "derived_dimension_count": row[7],
    }


def _git_output(
    arguments: Sequence[str],
    default: str,
) -> str:
    try:
        result = subprocess.run(
            ["git", *arguments],
            cwd=PROJECT_ROOT,
            check=True,
            text=True,
            capture_output=True,
        )
    except (OSError, subprocess.CalledProcessError):
        return default
    return result.stdout.strip()


def default_run_id(split: str) -> str:
    timestamp = datetime.now(timezone.utc).strftime(
        "%Y%m%dT%H%M%SZ"
    )
    return f"{split.lower()}-{timestamp}"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="官方题库批量运行与证据保存"
    )
    parser.add_argument(
        "--source",
        type=Path,
        default=DEFAULT_SOURCE,
    )
    parser.add_argument(
        "--split",
        choices=sorted(VALID_SPLITS),
        default="TRAIN",
    )
    parser.add_argument("--limit", type=int)
    parser.add_argument(
        "--question-id",
        action="append",
        dest="question_ids",
    )
    parser.add_argument("--run-id")
    parser.add_argument(
        "--runs-root",
        type=Path,
        default=DEFAULT_RUNS_ROOT,
    )
    parser.add_argument(
        "--resume",
        action=argparse.BooleanOptionalAction,
        default=True,
    )
    parser.add_argument(
        "--rerun-failures",
        action="store_true",
    )
    parser.add_argument(
        "--retries",
        type=int,
        default=1,
        help="瞬时网络异常的自动重试次数，默认1次。",
    )
    parser.add_argument(
        "--delay-seconds",
        type=float,
        default=0.0,
    )
    parser.add_argument(
        "--confirm-test-set",
        action="store_true",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    split = args.split.upper()

    if split == "TST" and not args.confirm_test_set:
        raise SystemExit(
            "TST仅用于阶段验收。确需运行时必须显式添加"
            "--confirm-test-set。"
        )

    questions = load_questions(args.source)
    selected = select_questions(
        questions=questions,
        split=split,
        limit=args.limit,
        question_ids=args.question_ids,
    )
    if not selected:
        raise SystemExit("没有选中任何题目。")

    run_id = args.run_id or default_run_id(split)
    paths = EvaluationPaths(
        project_root=PROJECT_ROOT,
        source=args.source,
        run_root=args.runs_root / run_id,
    )

    pipeline, runtime = build_runtime()
    manifest = build_manifest(
        run_id=run_id,
        split=split,
        source=args.source,
        selected_questions=selected,
        runtime=runtime,
    )

    print(f"运行编号：{run_id}")
    print(f"数据分区：{split}")
    print(f"选中题目：{len(selected)}")
    print(f"结果目录：{paths.run_root}")

    try:
        summary = execute_batch(
            pipeline=pipeline,
            questions=selected,
            paths=paths,
            manifest=manifest,
            resume=args.resume,
            rerun_failures=args.rerun_failures,
            retries=args.retries,
            delay_seconds=args.delay_seconds,
        )
    except KeyboardInterrupt:
        print(
            "\n运行已中断，当前进度已保存，可使用相同命令继续。",
            file=sys.stderr,
        )
        return 130

    print(
        json.dumps(
            summary,
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
